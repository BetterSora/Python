# -*- coding: utf-8 -*-
"""
Created on Thu Apr  6 09:46:09 2017

@author: Qin
"""

import openpyxl
import json

def txt2xlsx(filename):
    f = open(filename, 'r', encoding = 'utf-8')
    data = json.load(f, encoding = 'utf-8')
    f.close()
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    
    worksheet.title = 'city'
    
    for i in range(1, len(data) + 1):
        worksheet.cell(row = i, column = 1).value = i
        worksheet.cell(row = i, column = 2).value = data[str(i)]
    
    workbook.save('city.xlsx')

if __name__ == '__main__':
    txt2xlsx('city.txt')