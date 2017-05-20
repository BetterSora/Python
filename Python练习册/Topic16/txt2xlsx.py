# -*- coding: utf-8 -*-
"""
Created on Fri Apr  7 10:27:37 2017

@author: Qin
"""

import json
import openpyxl

def txt2xlsx(filename):
    f = open(filename, 'r', encoding = 'utf-8')
    data = json.load(f, encoding = 'utf-8')
    f.close()
    print(data)
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.title = 'numbers'
    
    for i in range(1, len(data) + 1):
        for j in range(1, len(data[i - 1]) + 1):
            if j == 1:
                worksheet.cell(row = i, column = 1).value = '['                
            worksheet.cell(row = i, column = j + 1).value = data[i - 1][j - 1]
        else:
            worksheet.cell(row = i, column = j + 2).value = ']'
    workbook.save('numbers.xlsx')
    
if __name__ == '__main__':
    txt2xlsx('numbers.txt')