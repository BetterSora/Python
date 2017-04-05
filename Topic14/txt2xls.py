# -*- coding: utf-8 -*-
"""
Created on Wed Apr  5 16:33:52 2017

@author: Qin
"""

import json
import openpyxl

def txt2xml(filename):
    f = open(filename, 'r', encoding = 'utf-8')
    #获得一个字典
    data = json.load(f, encoding = 'utf-8')
    f.close()
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    
    worksheet.title = 'student'
    for i in range(1, len(data) + 1):
        worksheet.cell(row = i, column = 1).value = i
        for j in range(len(data[str(i)])):
            worksheet.cell(row = i, column = j + 2).value = data[str(i)][j]
    workbook.save('student.xlsx')

if __name__ == '__main__':
    txt2xml('student.txt')